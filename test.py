# -*- coding: utf-8 -*-
import datetime
import openpyxl
import csv
import sys
import os
from openpyxl.styles import PatternFill     # 导入填充模块
from openpyxl.styles import Font

encode_type_alipay='gb18030'
encode_type_wechat='UTF-8'



except_alipay=["等待付款"]

color_green="92D050"
color_grey="E7E6E6"
color_yellow="FFFF00"
color_red="FF0000"
color_orange="FFC000"
color_blue="00B0F0"
color_none=None


def read_data(file_path):
    time_format=r"%Y-%m-%d %H:%M:%S"
    input_type=None
    data=[]
    if not os.path.exists(file_path):
        print("输入的文件不存在！")
        print("请检查该文件路径的正确性：%s"%file_path)
        sys.exit()

    try:
        with open(file_path, 'r',encoding='gb18030') as file:  #读取数据写入data
            reader = csv.reader(file)
            for row in reader:
                data.append(row)
    except:
        with open(file_path, 'r',encoding='UTF-8'  ) as file:
            reader = csv.reader(file)
            for row in reader:
                data.append(row)

    for i in range(len(data)):                                  #找到真正需要的部分
        if data[i]==[]:
            continue
        if input_type==None:
            if "支付宝" in data[i][0]:
                input_type="alipay"
            if "微信" in data[i][0]:
                input_type="wechat"
        if data[i][0]=="交易时间":
            break
    data=data[:i:-1]                                            #然后切除并从按日期从前往后排列

    if(input_type=="alipay"):
        i=0
        while(i<len(data)):
            try:
                data[i][0]= datetime.datetime.strptime(data[i][0], time_format) #修改为统一的时间格式
            except ValueError:
                time_format="%Y/%m/%d %H:%M"
                data[i][0]= datetime.datetime.strptime(data[i][0], time_format) #修改为统一的时间格式

            if(data[i][8] in except_alipay or float(data[i][6])==0):        #交易状态以及交易金额判断
                del data[i]
                continue
            if(data[i][8]=="交易关闭"):     #该两个条件满足时说明是退款
                if(data[i][5]=="不计收支"):
                    del data[i]
                    continue
            if(data[i][8]=="退款成功"):
                data[i][5]="退款"
            if(data[i][4][-4:]=="收益发放"):    #处理余额宝利息
                data[i][2]=""
                data[i][4]="余额宝利息"
                data[i][5]="收入"
            if(data[i][10][0]=="T"):            #处理淘宝商品
                data[i][2]=""
            if(data[i][2]=="美团"):             #处理美团商品
                data[i][2]=""
            if(data[i][4]=="余额宝-转出到银行卡"):
                data[i][4]="余额宝转出到"+data[i][2]
                data[i][5]="资金周转"
                data[i][2]=""
            if(data[i][4]=="余额宝-单次转入" ):
                data[i][5]="资金周转"
                data[i][2]=""

            data[i][4]=data[i][2]+data[i][4]    #讲其他无需进一步处理的交易记录的交易方和商品对象合在一起
            data[i][2]=""

            temp=[data[i][0],data[i][4],data[i][5],data[i][6],"支付宝",]
            data[i]=temp
            i+=1

    if(input_type=="wechat"):
        i=0
        while(i<len(data)):
            try:
                data[i][0]= datetime.datetime.strptime(data[i][0], time_format) #修改为统一的时间格式
            except:
                continue
            if(data[i][1]=="群收款" and data[i][4]=="收入"):
                data[i][2]="群收款_"+data[i][2]
                data[i][3]=""
                data[i][4]="群收款收入"

            if(data[i][1]=="微信红包"):
                data[i][2]="红包来自_"+data[i][2]
                data[i][3]=""
            
            if(data[i][1]=="微信红包（单发）"):
                data[i][3]=""

            if(data[i][1]=="微信红包-退款"):
                data[i][2]=""
                data[i][3]="微信红包-退款"

            if(data[i][1]=="商户消费"):
                data[i][3]=""
            
            if(data[i][1]=="零钱充值"):
                data[i][3]=data[i][1]+data[i][2][:-6]
                data[i][4]="资金周转"
                data[i][2]=""


            if(data[i][1]=="零钱提现"):
                data[i][3]=data[i][1]+data[i][2][:-6]
                data[i][2]=""
                data[i][4]="资金周转"
                service_fee=data[i+1][10][4:]
                data[i][5]="￥"+str(float(data[i][5][1:])-float(data[i][10][4:]))
                data.insert(i,[data[i][0],"提现手续费","支出",data[i][10][4:],"微信"])
                i=i+1

            data[i][3]=data[i][2]+data[i][3]    #讲其他无需进一步处理的交易记录的交易方和商品对象合在一起


            temp=[data[i][0],data[i][3],data[i][4],data[i][5][1:],"微信",]
            data[i]=temp
            i+=1
    
    return data

def format_data(data):
    data.sort()
    month_map={}
    for i in data:
        ym=str(i[0].year)+"-"+str(i[0].month).rjust(2,"0")  #将年月设置为XXXX-XX的形式
        if ym not in month_map.keys():                      #记录每个月出现记账的天数
            month_map[ym]=[i[0].day]                        #如果不存在某年月，就添加键
        else:
            month_map[ym].append(i[0].day)                  #若已存在就进行日的增加
        
        if i[-1]=="微信":
            if i[-3]=="收入":
                i.append(color_orange)
            elif i[-3]=="资金周转":
                i.append(color_green)
            elif i[-3]=="群收款收入":
                i.append(color_orange)
            else:
                i.append(color_grey)
        if i[-1]=="支付宝":
            if i[-3]=="收入":
                i.append(color_yellow)
            elif i[-3]=="资金周转":
                i.append(color_green)
            elif i[-3]=="退款":
                i.append(color_yellow)
            else:
                i.append(color_none)
    
    max_days=0
    main_ym=""
    for key in month_map.keys():
        days_in_month=len(set(month_map[key]))
        if max_days<days_in_month:
            max_days=days_in_month
            main_ym=key
    print("输入数据中，主要月份为:"+main_ym)
    return main_ym,data

def preProcess(main_ym):

    year  = main_ym[0:4]    
    month = main_ym[5:]
    month_firstday = datetime.datetime.strptime("%s-%s-1"%(year,month), '%Y-%m-%d')
    
    print("主要月份的最早起始日期为",end=" ")
    print(month_firstday.strftime("%Y-%m-%d"))
    month_firstday_weekday = month_firstday.weekday()+1
    if    month_firstday_weekday==1:
        pass
    elif  month_firstday_weekday<=4:
        month_firstday = month_firstday- datetime.timedelta(days=month_firstday_weekday-1)
    elif  month_firstday_weekday>=5:
        month_firstday = month_firstday+ datetime.timedelta(days=8-month_firstday_weekday)
    month_firstday_weekday = month_firstday.weekday()+1     #更新开始日期是星期几
    if month_firstday_weekday!=1:
        print("记账起始日期不为周一，请检查！") #检查用
    print("经计算：\n修改后表格的起始日期为",end=" ")
    print(month_firstday.strftime("%Y-%m-%d"))


    month =(int(month)+1)
    if month==13:
        year =str(int(year)+1)
        month=1
    month=str(month).rjust(2,"0")
    month_lastday = datetime.datetime.strptime("%s-%s-1"%(year,month), '%Y-%m-%d')
    month_lastday = month_lastday- datetime.timedelta(days=1)
    
    
    month_lastday_weekday = month_lastday.weekday()+1
    if    month_lastday_weekday==7:
        pass
    elif  month_lastday_weekday<=3:
        month_lastday = month_lastday- datetime.timedelta(days=month_lastday_weekday)
    elif  month_lastday_weekday>=4:
        month_lastday = month_lastday+ datetime.timedelta(days=7-month_lastday_weekday)
    month_lastday_weekday = month_lastday.weekday()+1       #更新结束日期是星期几
    if month_lastday_weekday!=7:
        print("记账结束日期不为周日，请检查！") #检查用
    print("修改后表格的结束日期为",end=" ")
    print(month_lastday.strftime("%Y-%m-%d"))
    week_num=int(((month_lastday-month_firstday).days+1)/7)
    print("总计%d周，共%d天"%(week_num,week_num*7))

    wb=openpyxl.Workbook()
    del wb["Sheet"]         #打开新的excel若为空，会自动创建名为"Sheet"的工作表
    for i in range(week_num):
        wb.create_sheet(index=i,title='第%s周'%(i+1))
        sheet=wb.worksheets[i]
        sheet['A1'].value='日期'
        sheet['B1'].value='项目'
        sheet['C1'].value='金额'
        sheet['D1'].value='可报销支出'
        sheet['E1'].value='其他支出'
        sheet['F1'].value='资金回收'
        sheet['G1'].value='收入'
        sheet['H1'].value='备注'
        
        sheet['A2'].value='开始时间'
        sheet['B2'].value='截止时间'

        sheet['A3'].value=month_firstday.strftime("%Y-%m-%d")+" 00:00:01"
        month_firstday+=datetime.timedelta(days=6)
        sheet['B3'].value=month_firstday.strftime("%Y-%m-%d")+" 23:59:59"
        month_firstday+=datetime.timedelta(days=1)

        sheet['A5'].value='收入'
        sheet['A6'].value='资金回收'
        sheet['A7'].value='支出'
        sheet['B5'].value='=SUM(G21:G:G)'
        sheet['B6'].value='=SUM(F21:F:F)'
        sheet['B7'].value='=SUM(C21:C:C)'

        sheet['A9' ].value='可报销支出'
        sheet['A10'].value='其他支出'
        sheet['A11'].value='实际个人支出'
        sheet['B9' ].value='=SUM(D21:D:D)'
        sheet['B10'].value='=SUM(E21:E:E)'
        sheet['B11'].value='=B7-B6-B9-B10'

        sheet['A20'].value='日期'
        sheet['B20'].value='项目'
        sheet['C20'].value='金额'
        sheet['D20'].value='可报销支出'
        sheet['E20'].value='其他支出'
        sheet['F20'].value='资金回收'
        sheet['G20'].value='收入'
        sheet['H20'].value='备注'

        sheet['A14'].value='银行卡余额'
        sheet['A15'].value='支付宝余额'
        sheet['A16'].value='微信余额'
        sheet['A17'].value='冻结金额'

        sheet['B13'].value='本周数据'
        sheet['C13'].value='上周数据'
        if(i!=0):
            sheet['C14'].value='=第'+str(i+1)+'周!B14'
            sheet['C15'].value='=第'+str(i+1)+'周!B15'
            sheet['C16'].value='=第'+str(i+1)+'周!B16'
            sheet['C17'].value='=第'+str(i+1)+'周!B17'


        sheet['D13'].value='支出'
        sheet['D14'].value='=C14-B14'
        sheet['D15'].value='=C15-B15'
        sheet['D16'].value='=C16-B16'
        sheet['D17'].value='=C17-B17'

        sheet['G11'].value='颜色'
        sheet['G12'].value='无填充'
        sheet['G13'].value='黄色';sheet['G13'].fill=PatternFill('solid',color_yellow)
        sheet['G14'].value='灰色';sheet['G14'].fill=PatternFill('solid',color_grey  )
        sheet['G15'].value='橙色';sheet['G15'].fill=PatternFill('solid',color_orange)
        sheet['G16'].value='浅绿';sheet['G16'].fill=PatternFill('solid',color_green )
        sheet['G17'].value='浅蓝';sheet['G17'].fill=PatternFill('solid',color_blue  )
        sheet['G18'].value='红色';sheet['G18'].fill=PatternFill('solid',color_red   )

        sheet['H11'].value='含义'
        sheet['H12'].value='支付宝支出'
        sheet['H13'].value='支付宝收入'
        sheet['H14'].value='微信支出'
        sheet['H15'].value='微信收入'
        sheet['H16'].value='资金周转'
        sheet['H17'].value='工资性收入'
        sheet['H18'].value='其他'

    wb.create_sheet(index=week_num,title='总结')
    sheet=wb.worksheets[week_num]
    line=1
    for i in range(week_num):
        sheet["A"+str(line)]="第%s周开销"%(i+1)
        sheet["B"+str(line)]="=第%s周!B7"%(i+1)
        line+=1

    for i in range(week_num):
        sheet["A"+str(line)]="第%s周可报销金额"%(i+1)
        sheet["B"+str(line)]="=第%s周!B9"%(i+1)
        line+=1

    for i in range(week_num):
        sheet["A"+str(line)]="第%s周群收款"%(i+1)
        sheet["B"+str(line)]="=第%s周!B6"%(i+1)
        line+=1

    for i in range(week_num):
        sheet["A"+str(line)]="第%s周其他支出"%(i+1)
        sheet["B"+str(line)]="=第%s周!B10"%(i+1)
        line+=1

    sheet["A"+str(line+1)]="总支出"
    sheet["A"+str(line+2)]="可报销支出"
    sheet["A"+str(line+3)]="群收款"
    sheet["A"+str(line+4)]="其他支出"
    sheet["B"+str(line+1)]="=SUM(B%s:B%s)"%(week_num*0+1,week_num*1)
    sheet["B"+str(line+2)]="=SUM(B%s:B%s)"%(week_num*1+1,week_num*2)
    sheet["B"+str(line+3)]="=SUM(B%s:B%s)"%(week_num*2+1,week_num*3)
    sheet["B"+str(line+4)]="=SUM(B%s:B%s)"%(week_num*3+1,week_num*4)

    sheet["A"+str(line+6)]="实际个人支出"
    sheet["B"+str(line+6)]="="+"B"+str(line+1)+"-B"+str(line+2)+"-B"+str(line+3)+"-B"+str(line+4)

    return wb,week_num

def write_excel_data(wb,data,week_num):
    startline=20
    old_week_day=-1
    sheet_num=0
    sheet=wb.worksheets[sheet_num]
    sheet_time_start=datetime.datetime.strptime(sheet["A3"].value, "%Y-%m-%d %H:%M:%S")
    sheet_time_over =datetime.datetime.strptime(sheet["B3"].value, "%Y-%m-%d %H:%M:%S")
    for i in data:
        if i[0]<sheet_time_start:           #若有时间不到记录范围，就判断下一条信息
            continue        
        if i[0]>sheet_time_over:            #若有时间超过了范围，就进入下一个sheet
            sheet_num+=1                    #同时初始化所有相关值
            if sheet_num>week_num-1:
                break
            
            sheet=wb.worksheets[sheet_num]
            startline=20
            old_week_day=-1
            sheet_time_start=datetime.datetime.strptime(sheet["A3"].value, "%Y-%m-%d %H:%M:%S")
            sheet_time_over =datetime.datetime.strptime(sheet["B3"].value, "%Y-%m-%d %H:%M:%S")
        
        now_week_day=i[0].weekday()+1       #如果出现日期变换，就要多加一行
        if now_week_day!=old_week_day:
            old_week_day=now_week_day
            startline+=1
            sheet['A'+str(startline)].value=i[0].strftime("%m月%d日")
        sheet['B'+str(startline)].value=i[1]
        if i[-1]!=None:
            sheet['B'+str(startline)].fill=PatternFill('solid', fgColor=i[-1])
        else:
            sheet['B'+str(startline)].fill=openpyxl.styles.PatternFill(fill_type=None)

        if i[2]=="收入" :
            sheet['G'+str(startline)].value=float(i[3])
        elif i[2]=="资金周转":
            sheet['B'+str(startline)].value=str(i[1]+i[3])
        elif i[2]=="群收款收入":
            sheet['F'+str(startline)].value=float(i[3])
        elif i[2]=="退款":
            sheet['F'+str(startline)].value=float(i[3])
        else:
            sheet['C'+str(startline)].value=float(i[3])
        startline+=1

def postProcess(wb,week_num):
    error=0.8
    for sheet_num in range(week_num):
        sheet=wb.worksheets[sheet_num]
        for row_num in range(1,sheet.max_row+1):          #最大行数
            sheet.row_dimensions[row_num-1].height=13.8
            for array_num in range(1,sheet.max_column+1):   #最大列数
                sheet.cell(row_num,array_num).font=Font(name="等线",size=11)
                if (row_num>20 and array_num>2):
                    sheet.cell(row_num,array_num).number_format = '\u00a5#,##0.00'    #设置了人民币的符号
                if (row_num>=2 and row_num<=11 and array_num==2):
                    sheet.cell(row_num,array_num).number_format = '\u00a5#,##0.00'    #设置了人民币的符号
                if (row_num>=14 and row_num<=17 and array_num>=2 and array_num<=4):
                    sheet.cell(row_num,array_num).number_format = '\u00a5#,##0.00'    #设置了人民币的符号
        
        

        sheet.column_dimensions["A"].width=20+error
        sheet.column_dimensions["B"].width=30+error
        sheet.column_dimensions["C"].width=20+error
        sheet.column_dimensions["D"].width=20+error
        sheet.column_dimensions["E"].width=20+error
        sheet.column_dimensions["F"].width=20+error
        sheet.column_dimensions["G"].width=20+error
        sheet.column_dimensions["H"].width=20+error
    
    sheet=wb.worksheets[sheet_num+1]
    sheet.column_dimensions["A"].width=20+error
    sheet.column_dimensions["B"].width=30+error
    for row_num in range(1,sheet.max_row+1):          #最大行数
            sheet.row_dimensions[row_num-1].height=13.8
            for array_num in range(1,sheet.max_column+1):   #最大列数
                sheet.cell(row_num,array_num).font=Font(name="等线",size=11)
                if (array_num==2):
                    sheet.cell(row_num,array_num).number_format = '\u00a5#,##0.00'    #设置了人民币的符号

def save_excel(wb,main_month):
    excenName_temp=main_month+"账单"        #尝试保存
    excenName     =excenName_temp
    i=0
    while(1):
        try:
            wb.save(excenName+'.xlsx')
        except:
            i=i+1
            excenName=excenName_temp+str(i)
        else:
            print("*************************")
            print("已保存为"+excenName+'.xlsx')
            break

def main():
    print("使用方法：")
    print("①命令行使用 python3 ZhangDanCheck.py <file path 1> <file path 2> ...")
    print("①运行程序后根据提示进行输入")

    data_total=[]
    if len(sys.argv)!=1:
        print(sys.argv)
        for i in sys.argv[1:]:
            data_total+=read_data(i)
    else:
        i=1
        file_path=[]
        print("请添加需要读取的excel表格路径，输入0来停止添加：")
        while 1:
            file_path_temp=input("第%d个待读取表格路径："%i)
            if file_path_temp=="0":
                break
            file_path.append(file_path_temp)
            i+=1
        for i in file_path:
            data_total+=read_data(i)



    main_month,data_total=format_data(data_total)

    wb,week_num=preProcess(main_month)         #写入初始框架

    write_excel_data(wb,data_total,week_num) #写入数据

    postProcess(wb,week_num)

    save_excel(wb,main_month,)
    

main()